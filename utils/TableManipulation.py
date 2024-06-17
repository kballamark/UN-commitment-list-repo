import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import difflib
import matplotlib.pyplot as plt
import pycountry
import numpy as np
from openpyxl.styles import PatternFill
from scipy.spatial import distance

class TableManipulation:
    def __init__(self):
        """
        Initializes the TableManipulation class. This class provides methods to read and manipulate
        data tables, specifically Excel files, including reformatting tables and adjusting column sizes.
        """
        self.input_dir = './input_tables/'
        self.output_dir = './output_tables/'

    def read_excel_file(self, file_name):
        """
        Reads an Excel file into a DataFrame.

        Parameters:
        - file_name: str, the name of the file to be read.

        Returns:
        - df: pandas.DataFrame, the DataFrame containing the data from the Excel file. Returns None if the file cannot be read.
        """
        file_path = self.input_dir + file_name
        try:
            df = pd.read_excel(file_path)
            return df
        except FileNotFoundError:
            print(f"File {file_name} not found in {self.input_dir}.")
            return None
        except Exception as e:
            print(f"An error occurred: {e}")
            return None
        
    def evaluate_sdgSix(self, dataframe):
        """
        Evaluates each row in the DataFrame based on its alignment with the SDG 6 (Clean Water and Sanitation) targets.
        This function iterates through each row of the provided DataFrame and counts the number of SDG 6 targets mentioned in the 'Targets' column. Each target matched adds a point to the row's score, which is then normalized based on the total number of SDG 6 targets. The final score for each row is added to the DataFrame as a new column named 'Evaluation_SDG'.

        Parameters:
        - dataframe (pandas.DataFrame): The DataFrame to be evaluated. It must contain a column named 'Targets' where the SDG 6 targets are listed.

        Returns:
        - pandas.DataFrame: The original DataFrame with an additional column 'Evaluation_SDG', containing the normalized scores representing the alignment of each row with SDG 6 targets.
        """

        # Define the targets of interest
        sdg_six_targets = ['6.1', '6.2', '6.3', '6.4', '6.5', '6.6', '6.a', '6.b']
        
        # Initialize a list to store the evaluation scores
        evaluation_scores = []
        
        # Iterate through each row in the dataframe
        for index, row in dataframe.iterrows():
            # Initialize the count for the current row
            count = 0
            
            # Check for each target in the sdg_six_targets list
            for target in sdg_six_targets:
                # If the target is found in the 'Targets' column of the current row, increment the count
                if target in str(row['Targets']):
                    count += 1
            
            # Calculate the score for the current row (1 point for each match, normalized by the total number of targets)
            score = count * (1/8)
            
            # Append the score to the evaluation_scores list
            evaluation_scores.append(score)
        
        # Add the evaluation scores as a new column in the dataframe
        dataframe['Evaluation_SDG'] = evaluation_scores
        
        # Return the modified dataframe
        return dataframe
    
    def evaluate_vulnerableGroups(self, dataframe, word_check, match_threshold=0.7):
        """
        Evaluates each commitment in the DataFrame based on its relevance to vulnerable groups, utilizing fuzzy matching to identify keywords within the text. 
        The evaluation score for each commitment is normalized based on the highest number of keyword occurrences, allowing for partial word matches determined by a match_threshold.

        Parameters:
        - dataframe (pd.DataFrame): The DataFrame containing commitments, with relevant columns to be analyzed for keywords.
        - word_check (list of str): A list of keywords associated with vulnerable groups.
        - match_threshold (float, optional): The similarity threshold for a word to be considered a match, default is 0.7 (70% match).

        Returns:
        - pd.DataFrame: The input DataFrame augmented with a new column 'Evaluation_vulnerables', containing normalized evaluation scores based on the presence of relevant keywords.
        """
        
        def get_match_count(text, word_list, threshold):
            count = 0
            for word in word_list:
                for text_word in text.split():
                    if difflib.SequenceMatcher(None, word.lower(), text_word.lower()).ratio() >= threshold:
                        count += 1
                        break  # Stop checking after the first match to avoid double-counting
            return count

        evaluation_scores = []
        max_occurrences = 0

        for index, row in dataframe.iterrows():
            combined_text = ' '.join([str(row[column]) for column in ['Beneficiaries', 'Description', 'Expected Impact', 'Partners']]).lower()
            count = get_match_count(combined_text, word_check, match_threshold)
            max_occurrences = max(max_occurrences, count)
            evaluation_scores.append(count)

        normalized_scores = [score / max_occurrences if max_occurrences > 0 else 0 for score in evaluation_scores]
        dataframe['Evaluation_vulnerables'] = normalized_scores

        return dataframe

    def evaluate_youthBeneficiaries(self, dataframe, word_check_list_explicit, word_check_list_implicit, match_threshold=0.7):
        """
        Evaluates each commitment based on youth involvement, with a focus on explicit and implicit keywords. 
        This function introduces fuzzy matching, allowing for partial matches between keywords and text, controlled by a match_threshold parameter.

        Parameters:
        - dataframe (pd.DataFrame): DataFrame containing commitments.
        - word_check_list_explicit (list of str): Keywords indicating explicit youth involvement.
        - word_check_list_implicit (list of str): Keywords indicating implicit youth involvement.
        - match_threshold (float, optional): The similarity threshold for a word to be considered a match, default is 0.7 (70% match).

        Returns:
        - pd.DataFrame: Modified DataFrame with 'Evaluation_youthBeneficiaries' scores, accounting for partial word matches.
        """
        
        def get_match_count(text, word_list, threshold):
            count = 0
            for word in word_list:
                for text_word in text.split():
                    if difflib.SequenceMatcher(None, word.lower(), text_word.lower()).ratio() >= threshold:
                        count += 1
                        break  # Stop checking after the first match to avoid double-counting
            return count

        scores_explicit = []
        scores_implicit = []
        max_occurrences_explicit = 0
        max_occurrences_implicit = 0

        for index, row in dataframe.iterrows():
            text = str(row['Beneficiaries']).lower()

            count_explicit = get_match_count(text, word_check_list_explicit, match_threshold)
            count_implicit = get_match_count(text, word_check_list_implicit, match_threshold)

            max_occurrences_explicit = max(max_occurrences_explicit, count_explicit)
            max_occurrences_implicit = max(max_occurrences_implicit, count_implicit)

            scores_explicit.append(count_explicit)
            scores_implicit.append(count_implicit)

        normalized_scores_explicit = [0.5 + 0.5 * (score / max_occurrences_explicit if max_occurrences_explicit > 0 else 0) for score in scores_explicit]
        normalized_scores_implicit = [0.5 * (score / max_occurrences_implicit if max_occurrences_implicit > 0 else 0) for score in scores_implicit]

        final_scores = [score_explicit if score_explicit > 0.5 else score_implicit for score_explicit, score_implicit in zip(normalized_scores_explicit, normalized_scores_implicit)]

        dataframe['Evaluation_youthBeneficiaries'] = final_scores

        return dataframe
    
    def evaluate_youthPartners(self, dataframe, word_check_list_explicit, word_check_list_implicit, match_threshold=0.7):
        """
        Evaluates each commitment based on youth involvement, with a focus on explicit and implicit keywords. 
        This function introduces fuzzy matching, allowing for partial matches between keywords and text, controlled by a match_threshold parameter.

        Parameters:
        - dataframe (pd.DataFrame): DataFrame containing commitments.
        - word_check_list_explicit (list of str): Keywords indicating explicit youth involvement.
        - word_check_list_implicit (list of str): Keywords indicating implicit youth involvement.
        - match_threshold (float, optional): The similarity threshold for a word to be considered a match, default is 0.7 (70% match).

        Returns:
        - pd.DataFrame: Modified DataFrame with 'Evaluation_youthPartners' scores, accounting for partial word matches.
        """
        
        def get_match_count(text, word_list, threshold):
            count = 0
            for word in word_list:
                for text_word in text.split():
                    if difflib.SequenceMatcher(None, word.lower(), text_word.lower()).ratio() >= threshold:
                        count += 1
                        break  # Stop checking after the first match to avoid double-counting
            return count

        scores_explicit = []
        scores_implicit = []
        max_occurrences_explicit = 0
        max_occurrences_implicit = 0

        for index, row in dataframe.iterrows():
            text = str(row['Partners']).lower()

            count_explicit = get_match_count(text, word_check_list_explicit, match_threshold)
            count_implicit = get_match_count(text, word_check_list_implicit, match_threshold)

            max_occurrences_explicit = max(max_occurrences_explicit, count_explicit)
            max_occurrences_implicit = max(max_occurrences_implicit, count_implicit)

            scores_explicit.append(count_explicit)
            scores_implicit.append(count_implicit)

        normalized_scores_explicit = [0.5 + 0.5 * (score / max_occurrences_explicit if max_occurrences_explicit > 0 else 0) for score in scores_explicit]
        normalized_scores_implicit = [0.5 * (score / max_occurrences_implicit if max_occurrences_implicit > 0 else 0) for score in scores_implicit]

        final_scores = [score_explicit if score_explicit > 0.5 else score_implicit for score_explicit, score_implicit in zip(normalized_scores_explicit, normalized_scores_implicit)]

        dataframe['Evaluation_youthPartners'] = final_scores

        return dataframe

    def reformat_table(self, df):
        """
        Reformats a DataFrame by resetting the 'ID' column to sequential numbers starting from 1.
        Additionally, this method adjusts the width of each column in the resulting Excel file based on
        the maximum length of the content in each column, enhancing the readability of the output.

        Parameters:
        - df: pandas.DataFrame, the DataFrame to be reformatted.

        Returns:
        - df: pandas.DataFrame, the modified DataFrame with sequential IDs and adjusted column widths.
        """
        if 'ID' in df.columns:
            df['ID'] = range(1, len(df) + 1)
        else:
            raise ValueError("DataFrame does not have an 'ID' column.")
        
        output_file_path = self.output_dir + 'reformatted_commitments_list.xlsx'
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length
        return df
    
    def filter_table(self, df, column_names, keywords, remove_non_matching=False, highlight_matches=False):
        """
        Filters a DataFrame based on the presence of keywords in specified columns, outputs and returns the IDs of matching rows,
        optionally removes non-matching rows, and highlights matching rows in an output Excel file.

        Parameters:
        - df: pandas.DataFrame, the DataFrame to be filtered.
        - column_names: set, a set containing strings as column names to search for keywords.
        - keywords: set, a set containing strings as keywords to search for in the specified columns.
        - remove_non_matching: bool, if True, removes rows that do not contain any of the keywords in any of the specified columns. Default is False.
        - highlight_matches: bool, if True, highlights matching rows in yellow in the output Excel file. Default is False.

        Outputs:
        - IDs of the rows matching the search criteria are printed to the terminal and returned as a list.

        Returns:
        - A tuple containing:
            - The modified DataFrame, which is saved to an Excel file named 'filtered_reformatted_commitments_list.xlsx' in the 'output_tables' folder, with auto-adjusted column widths.
            - A list of IDs corresponding to rows that match the search criteria.

        Notes:
        - The function assumes there is an 'ID' column in the DataFrame to extract the matching IDs.
        """
        import pandas as pd
        from openpyxl.utils import get_column_letter

        combined_mask = pd.Series([False] * df.shape[0])

        for column_name in column_names:
            if column_name in df.columns:
                mask = df[column_name].apply(lambda x: any(keyword in str(x).split() for keyword in keywords))
                combined_mask |= mask
            else:
                print(f"Warning: Column '{column_name}' does not exist in the DataFrame. Skipping this column.")

        matching_rows = df[combined_mask]
        total_rows = df.shape[0]

        matching_ids = []
        if 'ID' in df.columns:
            matching_ids = matching_rows['ID'].tolist()
            # Comment this out for now
            #print(f"IDs of matching rows: {matching_ids}")
        else:
            # Comment this out for now
            print("The DataFrame does not have an 'ID' column to output matching IDs.")

        print(f"Out of the {total_rows} commitments, {matching_rows.shape[0]} contains the keywords {', '.join(keywords)} in columns {', '.join(column_names)}.")

        if remove_non_matching:
            df = matching_rows

        output_file_path = './output_tables/filtered_reformatted_commitments_list.xlsx'
        
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length

        return df, matching_ids
    
    def highlight_table(self, df, ids):
        """
        Highlights the rows in a DataFrame that correspond to the provided IDs and exports the result to an Excel file.

        Parameters:
        - df: pandas.DataFrame, the DataFrame to be processed.
        - ids: list, a list of IDs corresponding to the rows that should be highlighted.

        Outputs:
        - An Excel file named 'highlighted_table.xlsx' in the 'output_tables' folder with the specified rows highlighted in yellow.
        """

        # Define the path for the output Excel file
        output_file_path = './output_tables/highlighted_table.xlsx'

        # Highlight matching rows in yellow
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Export to Excel
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # Apply column width adjustment
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length

            # Apply highlighting to rows with matching IDs
            if 'ID' in df.columns:
                for idx, row in enumerate(df.itertuples(), 1):  # Start enumeration at 1 due to Excel row indexing
                    if getattr(row, 'ID') in ids:
                        for cell in worksheet[f"{idx}:{idx}"]:
                            cell.fill = yellow_fill
            else:
                print("The DataFrame does not have an 'ID' column to match against provided IDs.")

        print(f"Highlighted table has been exported to {output_file_path}.")

    def evaluate_sectors(self, dataframe, entity_types):
        """
        Evaluates each row in the DataFrame based on the occurrence of specified entity types in the 'Entity Type' column. 
        The function assigns scores normalized by the total number of unique entity types provided and the occurrence frequencies.

        Parameters:
        - dataframe (pd.DataFrame): The DataFrame to be evaluated.
        - entity_types (list of str): A list of entity types to search for in the 'Entity Type' column.

        Returns:
        - pd.DataFrame: The input DataFrame augmented with a new column 'Evaluation_sectors', containing normalized scores.
        """
        
        # Initialize the score list and the maximum occurrence counter
        scores = []
        max_occurrences = 0

        # Iterate through each row in the DataFrame
        for index, row in dataframe.iterrows():
            # Initialize the count for the current row
            count = 0

            # Check for matches with the provided entity types
            for entity_type in entity_types:
                if entity_type.lower() == str(row['Entity Type']).lower().strip():
                    count += 1

            # Update the maximum occurrences if the current count is higher
            max_occurrences = max(max_occurrences, count)

            # Store the count for the current row
            scores.append(count)

        # Normalize the scores based on the maximum occurrences
        normalized_scores = [score / max_occurrences if max_occurrences > 0 else 0 for score in scores]

        # Add the normalized scores as a new column in the DataFrame
        dataframe['Evaluation_sectors'] = normalized_scores

        # Return the modified DataFrame
        return dataframe
    
    def evaluate_geography(self, dataframe, all_countries, match_threshold=0.9):
        """
        Evaluates each commitment in the DataFrame based on the number of countries mentioned in the 'Countries' column. 
        Special consideration is given to entries labeled as 'global', 'Global Action Plan', or 'European Commission', 
        which are directly assigned the maximum score of 1. The scores for other entries are normalized based on the 
        95th percentile of the country count distribution to mitigate the influence of outliers. Fuzzy matching is 
        employed to count country mentions, allowing for flexible identification of countries.

        Parameters:
        - dataframe (pd.DataFrame): The DataFrame to be evaluated, expected to contain a 'Countries' column.
        - all_countries (list of str): A comprehensive list of all country names used for fuzzy matching against the 'Countries' column.
        - match_threshold (float, optional): The similarity threshold for a fuzzy match between country names, defaulting to 0.9 (90% match).

        Returns:
        - pd.DataFrame: The input DataFrame augmented with a new 'Evaluation_geography' column containing normalized scores for each commitment. Scores range from 0 to 1, where a score of 1 indicates either a 'global' initiative or a number of countries mentioned at or above the 95th percentile of the distribution.

        Note:
        - Entries with 'global', 'Global Action Plan', or 'European Commission' in the 'Countries' column are considered to have a global impact and are automatically assigned the maximum score.
        - The normalization process uses the 95th percentile of the country count distribution to reduce the impact of outliers, providing a more balanced evaluation across all commitments.
        """

        scores = []
        for index, row in dataframe.iterrows():
            country_field = str(row['Countries']).lower()
            if 'global' in country_field or 'global action plan' in country_field or 'European Commission' in country_field:
                scores.append(1)
                continue

            countries_mentioned = country_field.split(',')
            count = 0
            for country in countries_mentioned:
                country = country.strip()
                for known_country in all_countries:
                    if difflib.SequenceMatcher(None, country, known_country.lower()).ratio() >= match_threshold:
                        count += 1
                        break

            scores.append(count)

        # Use the 95th percentile instead of the max value for normalization
        normalization_value = np.percentile(scores, 95)
        
        # Avoid division by zero
        if normalization_value > 0:
            normalized_scores = [min(score / normalization_value, 1) for score in scores]
        else:
            normalized_scores = [0 for _ in scores]
        
        dataframe['Evaluation_geography'] = normalized_scores
        return dataframe


    def plot_column_bar_chart(self, dataframe, column_name):
        """
        Plots a bar chart representing the frequencies of unique values in a specified column of a DataFrame,
        with x-axis tick labels rounded to two decimal places.

        Parameters:
        - dataframe (pd.DataFrame): The DataFrame containing the data to be plotted.
        - column_name (str): The name of the column for which the bar chart will be plotted.

        Returns:
        - None: The function directly displays the bar chart.
        """
        
        # Calculate the frequency of each unique value in the column
        value_counts = dataframe[column_name].value_counts().sort_index()

        # Plot the bar chart
        plt.figure(figsize=(10, 3))
        bar_plot = value_counts.plot(kind='bar', color='skyblue', edgecolor='black', alpha=0.7, width=1)
        plt.title(f'Frequency of Unique Values in {column_name}')
        plt.xlabel(column_name)
        plt.ylabel('Frequency')
        
        # Round x-axis tick labels to two decimal places and set them
        rounded_ticks = [f'{tick:.2f}' for tick in value_counts.index]
        bar_plot.set_xticklabels(rounded_ticks, rotation=45)

        plt.grid(axis='y', alpha=0.75)

        # Show the plot
        plt.show()


    def get_all_countries(self):
        """
        Generates a list of all country names using the pycountry library.

        Returns:
        - list of str: A list containing the official names of all countries.
        """
        
        country_names = [country.name for country in pycountry.countries]
        return country_names
    

    def calculate_weighted_sum_percentiles(self, dataframe, weights):
        """
        Calculates a weighted sum of evaluation metrics using provided percentage-based weights, converting scores to percentiles within each category to normalize impact.

        Parameters:
        - dataframe (pd.DataFrame): The DataFrame containing evaluation metrics.
        - weights (dict): A dictionary where keys are column names and values are the percentage-based weights for each column.

        Returns:
        - pd.DataFrame: The input DataFrame augmented with a 'summed_points' column representing the weighted sum, with all scores adjusted to percentiles.
        """
        
        # Convert percentage-based weights to decimal form
        decimal_weights = {key: value / 100 for key, value in weights.items()}

        # Select only the evaluation columns
        eval_columns = [col for col in dataframe.columns if col.startswith('Evaluation_')]
        
        # Convert scores within each column to percentiles
        for col in eval_columns:
            dataframe[col] = dataframe[col].rank(pct=True)
        
        # Calculate the weighted sum
        dataframe['summed_points'] = dataframe[eval_columns].apply(
            lambda row: sum(row[col] * decimal_weights[col] for col in eval_columns), axis=1
        )

        return dataframe

    def plot_scores_with_percentiles(self, dataframe, column_name):
        """
        Plots a scatter plot of evaluation scores against their percentiles in a specified column of a DataFrame,
        with the size of each dot proportional to the number of commitments at each percentile.

        Parameters:
        - dataframe (pd.DataFrame): The DataFrame containing the data to be plotted.
        - column_name (str): The name of the column for which the scatter plot will be created.

        Returns:
        - None: The function directly displays the scatter plot.
        """
        
        # Calculate percentiles for the evaluation scores
        scores = dataframe[column_name]
        percentiles = scores.rank(pct=True) * 100  # Convert ranks to percentiles

        # Calculate the size for each percentile based on frequency
        frequency = percentiles.value_counts()
        sizes = percentiles.map(frequency) * 10  # Scale up sizes for better visibility

        # Plot the scatter plot
        plt.figure(figsize=(10, 3))
        plt.scatter(percentiles, scores, s=sizes, color='blue', alpha=0.5)
        plt.title(f'Scatter Plot of Scores vs. Percentiles in {column_name}')
        plt.xlabel('Percentiles')
        plt.ylabel('Scores')
        plt.grid(True)

        # Set x and y limits to always show full scale
        plt.xlim(0, 100)
        plt.ylim(0, 1)

        # Show the plot
        plt.show()

    def topsis(self, dataframe, weights):
        # Normalize weights
        weight_sum = sum(weights.values())
        normalized_weights = {k: v / weight_sum for k, v in weights.items()}
        
        # Normalize the evaluation data
        eval_columns = [col for col in dataframe.columns if col.startswith('Evaluation_')]
        norm_data = dataframe[eval_columns].apply(lambda x: (x**2).sum()**0.5)
        norm_eval_data = dataframe[eval_columns].div(norm_data)
        
        # Prepare weighted data by multiplying each column by its corresponding weight
        for col in eval_columns:
            norm_eval_data[col] *= normalized_weights[col]
        
        # Determine ideal and negative-ideal solutions
        ideal_solution = norm_eval_data.max()
        negative_ideal_solution = norm_eval_data.min()
        
        # Calculate the distance to the ideal and negative-ideal solutions
        distance_to_ideal = norm_eval_data.apply(lambda x: distance.euclidean(x, ideal_solution), axis=1)
        distance_to_negative_ideal = norm_eval_data.apply(lambda x: distance.euclidean(x, negative_ideal_solution), axis=1)
        
        # Calculate the similarity to the ideal solution
        similarity_to_ideal = distance_to_negative_ideal / (distance_to_ideal + distance_to_negative_ideal)
        
        # Add the scores to the dataframe
        dataframe['TOPSIS_Score'] = similarity_to_ideal
        return dataframe.sort_values(by='TOPSIS_Score', ascending=False)

    def final_commitments(self, original_dataframe, top_ids, output_path='output_tables/final_commitments.xls'):
        """
        Filters and sorts the original DataFrame to include only the rows corresponding to specified IDs,
        sorted by their 'summed_points' in descending order, and exports the result to an Excel file.

        Parameters:
        - original_dataframe (pd.DataFrame): The original DataFrame loaded from the Excel table.
        - top_ids (list): A list of IDs that correspond to the top commitments.
        - output_path (str): The file path where the resulting Excel file will be saved.

        Returns:
        - None: The function saves the resulting DataFrame to an Excel file and does not return a value.
        """
        
        # Ensure the 'ID' column in original_dataframe is in the correct format (string or numeric) as in top_ids
        if isinstance(top_ids[0], str):
            original_dataframe['ID'] = original_dataframe['ID'].astype(str)
        else:
            original_dataframe['ID'] = original_dataframe['ID'].astype(int)

        # Filter the DataFrame to include only the rows with IDs in top_ids
        filtered_df = original_dataframe[original_dataframe['ID'].isin(top_ids)]
        
        # Sort the filtered DataFrame by 'summed_points' in descending order
        sorted_df = filtered_df.sort_values(by='summed_points', ascending=False)

        # Export the sorted DataFrame to an Excel file
        sorted_df.to_excel(output_path, index=False)

        import pandas as pd

    def rescore_dataframe(self, df, column_name):
        """
        Rescores values in the specified column of the dataframe.
        
        Parameters:
            df (pd.DataFrame): The dataframe to modify.
            column_name (str): The name of the column whose values are to be rescored.
        
        Returns:
            pd.DataFrame: The dataframe with the rescored values in the specified column.
        """
        # Define the mapping from original scores to new scores
        score_mapping = {1: 0, 2: 0.5, 3: 1}
        
        # Apply the mapping to the specified column
        df[column_name] = df[column_name].map(score_mapping)
        
        # Return the modified dataframe
        return df
    
    import pandas as pd

    def export_grouped_scenarios(self, original_df, sorted_ids, num_top, output_path='output_tables/final_commitments_topsis.xlsx'):
        """
        Exports an Excel file with selected data from multiple scenarios, each represented in separate columns,
        and highlights IDs that appear in multiple scenario columns.

        Parameters:
        - original_df (pd.DataFrame): The original DataFrame containing all data.
        - sorted_ids (dict): Dictionary with scenario names as keys and lists of IDs as values.
        - num_top (int): Number of top IDs to consider for each scenario.
        - output_path (str): Path to save the resulting Excel file.

        """
        # Create a new DataFrame to hold all the extracted data
        final_df = pd.DataFrame()

        # Iterate through each scenario to extract data
        for scenario, ids in sorted_ids.items():
            # Filter the top specified number of IDs
            filtered_df = original_df[original_df['ID'].isin(ids[:num_top])]

            # Ensure the order of IDs matches the sorted order
            filtered_df = filtered_df.set_index('ID').reindex(ids[:num_top]).reset_index()

            # Select only the necessary columns and rename them for clarity
            scenario_df = filtered_df[['ID', 'Title', 'Leadorg']]
            scenario_df.columns = [f'{col}-{scenario}' for col in scenario_df.columns]

            # Append to the final DataFrame
            final_df = pd.concat([final_df, scenario_df], axis=1)

        # Saving the DataFrame to Excel with highlighting
        writer = pd.ExcelWriter(output_path, engine='xlsxwriter')
        final_df.to_excel(writer, index=False, sheet_name='Top Commitments')

        # Get the workbook and the worksheet for formatting
        workbook  = writer.book
        worksheet = writer.sheets['Top Commitments']

        # Conditional formatting to highlight duplicate IDs across scenarios
        format = workbook.add_format({'bg_color': '#FFFF00'})  # Yellow background for duplicates

        for scenario in sorted_ids.keys():
            col_idx = pd.Index(final_df.columns).get_loc(f'ID-{scenario}')
            # Apply conditional formatting based on duplicates in the 'ID' columns
            worksheet.conditional_format(1, col_idx, num_top, col_idx, {
                'type': 'duplicate',
                'format': format
            })

        # Close the Pandas Excel writer and output the Excel file
        writer.save()















